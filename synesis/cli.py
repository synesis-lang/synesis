"""
cli.py - Interface de linha de comando do compilador Synesis

Proposito:
    Expor comandos de compilacao, validacao e inicializacao de projetos.
    Gerencia saida de diagnosticos e codigos de retorno.

Componentes principais:
    - main: grupo principal Click
    - compile/check/validate_template/init: comandos CLI

Dependencias criticas:
    - click: CLI
    - synesis.compiler: pipeline principal
    - synesis.parser/template_loader: validacao isolada

Exemplo de uso:
    synesis compile projeto.synp --json out.json --csv out_dir

Notas de implementacao:
    - Saidas usam formato arquivo:linha:coluna: [SEVERITY] mensagem.
    - --force permite exportacao mesmo com erros.

Gerado conforme: Especificacao Synesis v1.1
"""

from __future__ import annotations

import logging
import sys
import threading
import time
from pathlib import Path
from typing import Iterable

try:
    import click
except ImportError:
    raise ImportError(
        "click nao encontrado. CLI requer instalacao com: pip install synesis[cli]"
    )

from synesis import __version__ as VERSION
from synesis.compiler import CompilationStats, SynesisCompiler
from synesis.exporters.alpaca_export import export_alpaca
from synesis.exporters.csv_export import export_csv
from synesis.exporters.json_export import export_json
from synesis.exporters.xls_export import export_xls
from synesis.parser.lexer import SynesisSyntaxError, parse_file
from synesis.parser.template_loader import TemplateLoadError, load_template

# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _tty() -> bool:
    return hasattr(sys.stdout, "isatty") and sys.stdout.isatty()


def _c(text: str, **kwargs) -> str:
    return click.style(text, **kwargs) if _tty() else text


def _configure_logging(verbose: int, quiet: int) -> None:
    """Set root log level: -q → WARNING/ERROR, default → INFO, -v → DEBUG."""
    if quiet >= 2:
        level = logging.ERROR
    elif quiet == 1:
        level = logging.WARNING
    elif verbose >= 1:
        level = logging.DEBUG
    else:
        level = logging.INFO
    logging.basicConfig(level=level, format="[%(levelname)s] %(message)s")


_COPYRIGHT = "Copyright (c) 2011-2026 Christian Maciel de Britto"
_LICENSE   = "MIT (AGPL-3.0-only WITH Synesis-data-output-exception pending)"
_URL       = "https://github.com/synesis-lang/synesis"

_CREDITS_TEXT = f"""\
SYNESIS — Intellectual Genealogy
=================================
{_COPYRIGHT}
https://orcid.org/0000-0003-1431-3924

The Synesis compiler is the formal culmination of a research and
development trajectory spanning more than a decade. The following
prior works contributed to its architecture:

  BDM — Banco de Dados Multimodal (2011-2013)
  Master's dissertation, UFPR. DOI:10.13140/RG.2.2.10686.10563
  First definition of: sources, items, factors, relations,
  ontology, and knowledge graph as an integrated structure.

  SocioAtlas (2016-2018)
  Doctoral thesis, UFPR. DOI:10.13140/RG.2.2.26449.17760
  Integration of annotations, audit trails, Zotero,
  georeferenced data, and knowledge graphs.

  DSAP annotation pipeline (2019-2020)
  Professional consultancy, environmental sector.
  First professional validation of the audit trail:
  corpus → item → summary → theme → score.

  SocioAtlas para Google Sheets (2022)
  Independent development, Google Workspace Marketplace.
  Collaboration and portability; first attempt at systematic
  theological study within the same framework.

  DGT.7 pipeline (2024)
  Independent development.
  Text-file knowledge representation; exposed the need
  for formal, readable, validatable syntax.

Full history:
  https://synesis-lang.github.io/synesis-docs/pt/explanation/sobre.html
  https://synesis-lang.github.io/synesis-docs/en/explanation/about.html
"""


def _build_main_help() -> str:
    title = _c("SYNESIS COMPILER", fg="green", bold=True) + f" (v{VERSION})"
    copyright_line = _c(_COPYRIGHT, fg="bright_black")
    url_line       = _c(_URL, fg="bright_black")
    desc = "Semantic compiler for knowledge engineering."
    license_line = _c(f"Licensed under {_LICENSE}.", fg="bright_black")
    usage = _c("Usage:", fg="yellow", bold=True) + " synesis [COMMAND] [OPTIONS]"

    groups = [
        ("Project Management", [
            ("init",              "Creates the minimal structure for a new project"),
        ]),
        ("Compilation & Export", [
            ("compile",           "Compiles one project (or links several) and generates artifacts (JSON, CSV, XLS, Alpaca)"),
        ]),
        ("Validation & Debugging", [
            ("check",             "Validates the syntax and integrity of a single file"),
            ("validate-template", "Verifies the structure and consistency of a template file"),
        ]),
    ]

    opt_rows = [
        ("-v, --verbose",  "Increase log verbosity (DEBUG). Repeatable."),
        ("-q, --quiet",    "Decrease log verbosity (-q WARNING, -qq ERROR). Repeatable."),
        ("--version",      "Show version and exit"),
        ("--credits",      "Show intellectual genealogy and prior works"),
        ("--help",         "Show this message and exit"),
    ]

    cmd_names_len = max(len(name) for _, rows in groups for name, _ in rows)
    opt_names_len = max(len(name) for name, _ in opt_rows)
    col = max(cmd_names_len, opt_names_len) + 2

    options = _c("Global Options:", fg="yellow", bold=True) + "\n" + "\n".join(
        f"  {_c(name.ljust(col), fg='cyan')}  {desc_}"
        for name, desc_ in opt_rows
    )

    def _render_group(label, rows):
        lines = [_c("  " + label, fg="yellow", bold=True)]
        for name, desc_ in rows:
            lines.append(f"    {_c(name.ljust(col), fg='green', bold=True)}  {desc_}")
        return "\n".join(lines)

    commands = _c("Commands:", fg="yellow", bold=True) + "\n\n" + "\n\n".join(
        _render_group(label, rows) for label, rows in groups
    )

    hint = _c(
        "Run 'synesis COMMAND --help' for specific parameters, output formats, and examples.",
        fg="bright_black",
    )

    return "\n\n".join([title, copyright_line, url_line, desc, license_line, usage, options, commands, hint]) + "\n"


class _SynesisCommand(click.Command):
    def format_epilog(self, ctx, formatter):
        if self.epilog:
            formatter.write("\n")
            for line in self.epilog.splitlines():
                formatter.write(line + "\n")


class _SynesisGroup(click.Group):
    command_class = _SynesisCommand

    def format_help(self, ctx, formatter):
        pass

    def get_help(self, ctx):
        out = _build_main_help()
        if hasattr(sys.stdout, "buffer"):
            sys.stdout.buffer.write(out.encode("utf-8"))
            sys.stdout.buffer.flush()
            raise SystemExit(0)
        return out


def _ex(*lines: str) -> str:
    import re
    out = [_c("Examples:", fg="yellow", bold=True)]
    for line in lines:
        stripped = line.lstrip()
        indent = line[: len(line) - len(stripped)]
        if stripped.startswith("#"):
            out.append(indent + _c(stripped, fg="bright_black"))
        else:
            tokens = re.split(r"(\s+)", stripped)
            result = []
            for tok in tokens:
                if tok == "synesis":
                    result.append(_c(tok, fg="green", bold=True))
                elif re.match(r"^--[\w-]+=?", tok):
                    result.append(_c(tok, fg="cyan"))
                elif tok in ("compile", "check", "validate-template", "init"):
                    result.append(_c(tok, fg="green"))
                else:
                    result.append(tok)
            out.append(indent + "".join(result))
    return "\n".join(out)


_EPILOG_COMPILE = _ex(
    "  # Compile and export to JSON",
    "  synesis compile project.synp --json output.json",
    "",
    "  # Compile and export to CSV directory",
    "  synesis compile project.synp --csv output_csv/",
    "",
    "  # Compile and export to Excel",
    "  synesis compile project.synp --xls output.xlsx",
    "",
    "  # Export Alpaca JSONL for LLM fine-tuning",
    "  synesis compile project.synp --alpaca dataset.jsonl",
    "",
    "  # Combine multiple export formats",
    "  synesis compile project.synp --json out.json --csv out_csv/ --alpaca dataset.jsonl",
    "",
    "  # Show compilation statistics",
    "  synesis compile project.synp --stats",
    "",
    "  # Link two or more projects: fields declared with IDENTIFIES/REFERS TO",
    "  # in the templates are resolved into edges across the aggregate.",
    "  synesis compile lattes.synp abstracts.synp",
    "",
    "  # Link N projects and export the aggregate (JSON v3.1 + links.csv)",
    "  synesis compile lattes.synp abstracts.synp --json export.json --csv out_csv/",
    "",
    "  # Show per-member and aggregate statistics (sources, items, edges, orphans)",
    "  synesis compile lattes.synp abstracts.synp --stats",
    "",
    "  # A single project always compiles isolated — the linking path only",
    "  # activates when 2+ .synp files are given. Isolated compilation of a",
    "  # project that declares REFERS TO emits an informational note (never a",
    "  # warning); compile it together with the project that owns IDENTIFIES",
    "  # for that entity to resolve the edges.",
    "",
    "  # Export the linked package as spreadsheets: --xls names a DIRECTORY here,",
    "  # with one .xlsx per member plus links.xlsx (resolved edges, with a readable",
    "  # label from each side). SOURCE FIELDS differ across members, so there is no",
    "  # single coherent table — the package is tabular per member.",
    "  synesis compile lattes.synp abstracts.synp --xls quinto_andar/",
    "",
    "  # --alpaca still has no per-member exporter in the linking path: the CLI",
    "  # warns instead of silently skipping it — run that export per project.",
)

_EPILOG_CHECK = _ex(
    "  # Validate a single annotations file",
    "  synesis check annotations.syn",
    "",
    "  # Validate a template file syntax",
    "  synesis check template.synt",
)

_EPILOG_VALIDATE_TEMPLATE = _ex(
    "  # Verify a template before running compile",
    "  synesis validate-template template.synt",
)

_EPILOG_INIT = _ex(
    "  # Initialize a new project in the current directory",
    "  synesis init",
)


# ---------------------------------------------------------------------------
# Spinner
# ---------------------------------------------------------------------------

class _Spinner:
    """Spinner animado para etapas do pipeline. Desativa-se se stdout nao for TTY."""

    _FRAMES = ("⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏")

    def __init__(self) -> None:
        self._active = False
        self._thread: threading.Thread | None = None
        self._label = ""
        self._is_tty = sys.stderr.isatty()

    def start(self, label: str) -> None:
        self._label = label
        if not self._is_tty:
            sys.stderr.write(f"  {label}...\n")
            sys.stderr.flush()
            return
        self._active = True
        self._t0 = time.monotonic()
        self._thread = threading.Thread(target=self._spin, daemon=True)
        self._thread.start()

    def _spin(self) -> None:
        i = 0
        while self._active:
            frame = self._FRAMES[i % len(self._FRAMES)]
            sys.stderr.write(f"\r  {frame} {self._label}...")
            sys.stderr.flush()
            time.sleep(0.08)
            i += 1

    def done(self, suffix: str = "") -> None:
        if not self._is_tty:
            return
        self._active = False
        if self._thread:
            self._thread.join()
        elapsed = time.monotonic() - self._t0
        elapsed_str = click.style(f"({elapsed:.1f}s)", fg="bright_black")
        check = click.style("✔", fg="green", bold=True)
        suffix_str = f"  {suffix}" if suffix else ""
        sys.stderr.write(f"\r  {check} {self._label}{suffix_str}  {elapsed_str}\n")
        sys.stderr.flush()

    def fail(self) -> None:
        if not self._is_tty:
            return
        self._active = False
        if self._thread:
            self._thread.join()
        cross = click.style("✖", fg="red", bold=True)
        sys.stderr.write(f"\r  {cross} {self._label}\n")
        sys.stderr.flush()




def _version_callback(ctx: click.Context, _param: click.Parameter, value: bool) -> None:
    if not value or ctx.resilient_parsing:
        return
    lines = [
        f"synesis {VERSION}",
        _COPYRIGHT,
        f"License: {_LICENSE}",
        _URL,
    ]
    click.echo("\n".join(lines))
    ctx.exit()


def _credits_callback(ctx: click.Context, _param: click.Parameter, value: bool) -> None:
    if not value or ctx.resilient_parsing:
        return
    click.echo(_CREDITS_TEXT, nl=False)
    ctx.exit()


@click.group(cls=_SynesisGroup, invoke_without_command=True)
@click.option("--version", is_flag=True, is_eager=True, expose_value=False,
              callback=_version_callback, help="Show version and exit.")
@click.option("--credits", is_flag=True, is_eager=True, expose_value=False,
              callback=_credits_callback, help="Show intellectual genealogy and prior works.")
@click.option("-v", "--verbose", count=True, default=0,
              help="Increase log verbosity (-v for DEBUG). Repeatable.")
@click.option("-q", "--quiet", count=True, default=0,
              help="Decrease log verbosity (-q for WARNING, -qq for ERROR). Repeatable.")
@click.pass_context
def main(ctx: click.Context, verbose: int, quiet: int) -> None:
    """Compilador semântico para validação e consolidação de conhecimento."""
    _configure_logging(verbose, quiet)
    if ctx.invoked_subcommand is None:
        out = _build_main_help()
        if hasattr(sys.stdout, "buffer"):
            sys.stdout.buffer.write(out.encode("utf-8"))
            sys.stdout.buffer.flush()
        else:
            click.echo(out)


@main.command(cls=_SynesisCommand, epilog=_EPILOG_COMPILE)
@click.argument("projects", type=click.Path(exists=True), nargs=-1, required=True)
@click.option("--json", "json_path", type=click.Path(), help="Export canonical JSON v3.0")
@click.option("--csv", "csv_dir", type=click.Path(), help="Export CSV tables to directory")
@click.option("--xls", "xls_path", type=click.Path(), help="Export Excel workbook (.xlsx)")
@click.option("--alpaca", "alpaca_path", type=click.Path(), help="Export Alpaca JSONL for LLM fine-tuning")
@click.option("--strict", is_flag=True, help="Treat warnings as errors")
@click.option("--stats", is_flag=True, help="Show compilation statistics")
@click.option("--force", is_flag=True, help="Generate artifacts even with errors")
def compile(projects: tuple[str, ...], json_path: str | None, csv_dir: str | None, xls_path: str | None, alpaca_path: str | None, strict: bool, stats: bool, force: bool) -> None:
    """Compile one Synesis project, or link several (compile p1 p2 ...)."""
    if len(projects) == 1:
        _compile_single(projects[0], json_path, csv_dir, xls_path, alpaca_path, strict, stats, force)
    else:
        _link_projects(projects, json_path, csv_dir, xls_path, alpaca_path, strict, stats, force)


def _compile_single(project: str, json_path: str | None, csv_dir: str | None, xls_path: str | None, alpaca_path: str | None, strict: bool, stats: bool, force: bool) -> None:
    """Compile a single Synesis project (unchanged legacy path)."""
    click.echo(click.style(f"SYNESIS v{VERSION}", bold=True) + "  Compile seu pensamento.")

    spinner = _Spinner()
    project_path = Path(project)
    project_dir = project_path.parent

    try:
        from synesis.ast.results import MalformedBibliographyEntry, ValidationResult
        from synesis.parser.template_loader import validate_template

        compiler = SynesisCompiler(project_path)

        # Etapa 1: projeto + template + bibliografia
        spinner.start("Lendo projeto e template")
        project_node, project_validation = compiler.parse_project()
        project_validation_structure = compiler.validate_project_structure(project_node)
        bib_validation = compiler._check_bibliography_file(project_node)
        bib_format_validation = compiler._check_bibliography_format(project_node)
        template, template_load_result = compiler._safe_load_template(project_node)
        if template is None:
            spinner.fail()
            result_early = ValidationResult()
            compiler._merge(result_early, project_validation)
            compiler._merge(result_early, project_validation_structure)
            compiler._merge(result_early, template_load_result)
            compiler._merge(result_early, bib_validation)
            compiler._merge(result_early, bib_format_validation)
            _print_diagnostics(result_early.errors, "ERROR", project_dir)
            raise SystemExit(1)
        template_validation = validate_template(template)
        bibliography = compiler.load_bibliography(project_node)
        dataset, dataset_load_result = compiler.load_dataset_index(project_node, template)
        spinner.done()

        # Etapa 2: ontologia
        spinner.start("Carregando ontologia")
        ontologies, ontology_load_result = compiler.parse_ontologies(project_node)
        spinner.done(f"{len(ontologies):,}".replace(",", ".") + " conceitos")

        # Etapa 3: anotacoes (etapa mais lenta — paralela)
        spinner.start("Lendo anotacoes")
        sources, items, annotations_load_result = compiler.parse_annotations(project_node)
        spinner.done(
            f"{len(sources):,}".replace(",", ".") + " sources, "
            f"{len(items):,}".replace(",", ".") + " items"
        )

        # Etapa 4: validacao semantica
        spinner.start("Validando")
        norm_cache: dict = {}
        malformed_keys = {
            e.entry_key.lower() for e in bib_format_validation.errors
            if isinstance(e, MalformedBibliographyEntry)
        }
        validation_result = compiler.validate_all(
            project=project_node,
            template=template,
            bibliography=bibliography,
            sources=sources,
            items=items,
            ontologies=ontologies,
            norm_cache=norm_cache,
            malformed_bib_keys=malformed_keys,
            dataset=dataset,
        )
        compiler._merge(validation_result, project_validation)
        compiler._merge(validation_result, project_validation_structure)
        compiler._merge(validation_result, template_validation)
        compiler._merge(validation_result, bib_validation)
        compiler._merge(validation_result, bib_format_validation)
        compiler._merge(validation_result, dataset_load_result)
        compiler._merge(validation_result, ontology_load_result)
        compiler._merge(validation_result, annotations_load_result)
        n_errors = len(validation_result.errors)
        n_warnings = len(validation_result.warnings)
        if n_errors:
            spinner.done(click.style(f"{n_errors} erro(s)", fg="red"))
        elif n_warnings:
            spinner.done(click.style(f"{n_warnings} aviso(s)", fg="yellow"))
        else:
            spinner.done("sem erros")

        # Etapa 5: vinculacao
        spinner.start("Vinculando")
        linked_project = compiler.link_all(
            project=project_node,
            template=template,
            sources=sources,
            items=items,
            ontologies=ontologies,
            validation_result=validation_result,
            norm_cache=norm_cache,
        )
        spinner.done()

        result_stats = compiler._compute_stats(linked_project, sources, items, ontologies)

        click.echo("")
        _print_diagnostics(validation_result.errors, "ERROR", project_dir)
        _print_diagnostics(validation_result.warnings, "WARNING", project_dir)
        _print_diagnostics(validation_result.info, "INFO", project_dir)

        if stats:
            click.echo("")
            _print_stats(result_stats)

        has_errors = validation_result.has_errors()
        has_warnings = validation_result.has_warnings()
        exit_code = 1 if has_errors or (strict and has_warnings) else 0

        if (force or exit_code == 0) and linked_project:
            if json_path:
                export_json(linked_project, Path(json_path), template, bibliography, dataset)
            if csv_dir:
                export_csv(linked_project, template, Path(csv_dir), bibliography, dataset)
            if xls_path:
                export_xls(linked_project, template, Path(xls_path), bibliography, dataset)
            if alpaca_path:
                export_alpaca(linked_project, Path(alpaca_path), template, bibliography)

        raise SystemExit(exit_code)

    except SystemExit:
        raise
    except SynesisSyntaxError as exc:
        spinner.fail()
        click.echo(click.style(str(exc), fg="red"), err=True)
        raise SystemExit(1)
    except Exception as exc:
        spinner.fail()
        click.echo(click.style(f"erro: Falha inesperada durante compilacao: {exc}", fg="red"), err=True)
        if click.get_current_context().obj and click.get_current_context().obj.get("debug"):
            raise
        raise SystemExit(1)


def _member_label_field(template) -> str | None:
    """Campo SCOPE SOURCE que serve de rotulo humano do membro no links.xlsx.

    Uma tabela de arestas so com bibrefs e chaves estrangeiras exige abrir os
    outros arquivos para ser lida. Escolhe-se o primeiro campo TEXT de SCOPE
    SOURCE que NAO seja a propria chave do link (IDENTIFIES/REFERS TO) — em
    lattes.synt isso da `nome`; em abstracts.synt, `description`. Heuristica
    deliberadamente simples: quando erra, degrada para coluna vazia, nunca
    para dado errado.
    """
    from synesis.ast.nodes import FieldType, Scope

    if template is None:
        return None
    for name, spec in template.field_specs.items():
        if spec.scope != Scope.SOURCE:
            continue
        if getattr(spec, "identifies", None) or getattr(spec, "refers_to", None):
            continue
        if spec.type == FieldType.TEXT:
            return name
    return None


def _sheet_rows(ws) -> list[tuple]:
    """Conteudo de uma worksheet como lista de tuplas (cabecalho incluido)."""
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def _shares_ontology(project_path: Path) -> bool:
    """True se o .synp declara INCLUDE SHARED ONTOLOGY.

    O `shared` do IncludeNode e a declaracao explicita do autor de que aquela
    ontologia e comum a varios projetos (D13 autoriza o escape de pasta so
    nesse caso). E o sinal correto para fundir as abas de ontologia no
    workbook unificado — sem ele, duas ontologias iguais por coincidencia
    seriam fundidas e a coincidencia viraria contrato silencioso.
    """
    try:
        node, _ = SynesisCompiler(project_path).parse_project()
    except Exception:  # noqa: BLE001 - projeto ja compilou antes; aqui e best-effort
        return False
    return any(
        inc.include_type.upper() == "ONTOLOGY" and getattr(inc, "shared", False)
        for inc in node.includes
    )


def _collapse_identical_sheets(wb, aliases: list[str], bases: tuple[str, ...]) -> list[str]:
    """Funde abas identicas entre TODOS os membros que compartilham ontologia.

    `INCLUDE SHARED ONTOLOGY` faz os membros carregarem o mesmo `.syno`, entao
    `ontologies`/`topics` saem repetidas uma vez por membro — no corpus real
    isso e metade das celulas do arquivo. Fundidas, viram uma aba sem prefixo
    (`ontologies`) valida para todos.

    Duas condicoes, ambas necessarias:
      1. todos os membros declaram INCLUDE SHARED ONTOLOGY — a intencao do
         autor de que a ontologia seja comum;
      2. o conteudo exportado coincide de fato.
    A (1) sozinha nao basta porque a ontologia pode ter sido editada entre
    compilacoes; a (2) sozinha fundiria por coincidencia dois projetos que
    apenas por acaso tem a mesma ontologia. Divergindo, as abas por membro
    permanecem — redundancia visivel e melhor que fusao que esconde diferenca.
    """
    collapsed: list[str] = []
    if len(aliases) < 2:
        return collapsed

    for base in bases:
        names = [f"{alias}_{base}" for alias in aliases]
        if not all(n in wb.sheetnames for n in names):
            continue

        reference = _sheet_rows(wb[names[0]])
        if not all(_sheet_rows(wb[n]) == reference for n in names[1:]):
            continue  # divergem: mantem uma aba por membro

        # Renomeia a primeira para o nome sem prefixo e descarta as demais.
        wb[names[0]].title = base
        for n in names[1:]:
            wb.remove(wb[n])
        collapsed.append(base)

    # As abas fundidas valem para todos os membros, entao nao pertencem ao
    # bloco de nenhum deles: vao para o fim, logo antes de `links` (que a
    # chamadora cria depois). Sem isto ficariam no meio das abas do primeiro
    # membro, sugerindo que sao dele.
    for base in collapsed:
        wb.move_sheet(base, offset=len(wb.sheetnames) - wb.sheetnames.index(base) - 1)

    return collapsed


def _export_unified_workbook(link_result, members, member_results, path: Path) -> list[str]:
    """Um unico .xlsx com as abas de todos os membros, prefixadas por alias.

    Cada membro mantem seu proprio esquema numa aba propria
    (`lattes_sources`, `abstracts_sources`, ...) — nao se tenta fundir colunas
    incompativeis (§6). O que a unificacao resolve e a circulacao (um anexo em
    vez de N) e a ligacao: dentro de um workbook, referencias entre abas sao
    estaveis, ao contrario de vinculos externos entre arquivos, que dependem
    de caminho absoluto e quebram ao mover o arquivo.

    Abas identicas entre todos os membros (tipicamente as da ontologia
    compartilhada) sao fundidas numa so — ver _collapse_identical_sheets.

    A aba `links` fecha o circuito: alem dos bibrefs qualificados, traz o
    rotulo legivel de cada lado como VALOR (nao formula) — legivel por humano e
    por pandas/openpyxl, que nao calculam formulas.

    Returns:
        Nomes das abas fundidas (vazio quando nenhuma coincidiu).
    """
    from synesis.exporters.xls_export import build_xls_workbook

    wb = None
    for m in members:
        res = member_results[m.alias]
        wb = build_xls_workbook(
            res.linked_project, res.template,
            res.bibliography, res.dataset,
            workbook=wb, prefix=m.alias,
        )

    # Candidatas a fusao: abas cujo conteudo deriva da ontologia, nao das
    # anotacoes do membro. `chains`/`code_frequency`/`items`/`sources` sao
    # dados proprios de cada membro e nunca coincidem — nem se tenta.
    # So quando TODOS os membros declaram INCLUDE SHARED ONTOLOGY.
    collapsed: list[str] = []
    if all(_shares_ontology(m.path) for m in members):
        collapsed = _collapse_identical_sheets(
            wb, [m.alias for m in members], ("ontologies", "topics")
        )

    _fill_links_sheet(link_result, members, member_results, wb.create_sheet("links"))
    wb.save(path)
    return collapsed


def _links_rows(link_result, members, member_results):
    """Linhas da tabela de arestas: cabecalho + arestas resolvidas + orfaos."""
    from synesis.exporters._helpers import _get_source_field_value

    label_field = {m.alias: _member_label_field(m.template) for m in members}

    def _label(alias: str, qualified_bibref: str) -> str:
        field = label_field.get(alias)
        res = member_results.get(alias)
        if not field or res is None or res.linked_project is None:
            return ""
        bibref = qualified_bibref.split(":@", 1)[-1]
        source = res.linked_project.sources.get(bibref.lower())
        if source is None:
            source = res.linked_project.sources.get(bibref)
        if source is None:
            return ""
        value = _get_source_field_value(
            source, field, res.template, res.bibliography, res.dataset
        )
        text = "" if value is None else str(value)
        return text[:200]

    yield [
        "entity", "value",
        "from_member", "from_bibref", "from_label",
        "to_member", "to_bibref", "to_label",
    ]
    for e in link_result.edges:
        yield [
            e.entity, e.value,
            e.from_member, e.from_bibref, _label(e.from_member, e.from_bibref),
            e.to_member, e.to_bibref, _label(e.to_member, e.to_bibref),
        ]
    for ent, val, mem in link_result.orphans:
        yield [ent, val, mem, "", "", "(orphan)", "", ""]


def _fill_links_sheet(link_result, members, member_results, ws) -> None:
    """Preenche uma worksheet `links` e ajusta a largura das colunas."""
    for row in _links_rows(link_result, members, member_results):
        ws.append(row)
    for column in ws.columns:
        width = max((len(str(c.value)) for c in column if c.value), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(width + 2, 50)


def _write_links_xlsx(link_result, members, member_results, path: Path) -> None:
    """Escreve links.xlsx: arestas resolvidas + orfaos, com rotulo de cada lado.

    Espelha o links.csv, mas acrescenta as colunas `from_label`/`to_label` —
    sem elas a tabela e uma lista de chaves estrangeiras, ilegivel sem abrir os
    .xlsx dos membros.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "links"
    _fill_links_sheet(link_result, members, member_results, ws)
    wb.save(path)


def _link_projects(projects: tuple[str, ...], json_path: str | None, csv_dir: str | None, xls_path: str | None, alpaca_path: str | None, strict: bool, stats: bool, force: bool) -> None:
    """Link step: compila N projetos isolados e resolve IDENTIFIES/REFERS TO entre eles.

    Modelo do linker C/C++ (D2): a agregacao e modo de CLI, nunca de LSP. Cada
    membro compila isolado; este passo resolve os simbolos entre unidades.
    """
    import csv as _csv
    import json as _json

    from synesis.semantic.link_step import Member, link_members

    click.echo(click.style(f"SYNESIS v{VERSION}", bold=True) + "  Linkando projetos.")

    members: list[Member] = []
    member_stats: list[tuple[str, CompilationStats, set]] = []
    member_results: dict[str, object] = {}  # alias -> CompilationResult (para export por membro)
    member_errors = False
    for proj in projects:
        proj_path = Path(proj)
        alias = proj_path.stem
        click.echo(f"  compilando {click.style(alias, bold=True)} …")
        result = SynesisCompiler(proj_path).compile()
        if result.has_errors():
            member_errors = True
            _print_diagnostics(result.validation_result.errors, "ERROR", proj_path.parent)
            continue
        if result.linked_project is None or result.template is None:
            member_errors = True
            click.echo(click.style(f"erro: `{alias}` nao produziu artefato linkavel.", fg="red"), err=True)
            continue
        members.append(Member(
            alias=alias,
            template=result.template,
            sources=result.linked_project.sources,
            path=proj_path,
            bibliography=result.bibliography or {},
            dataset=result.dataset or {},
        ))
        # Conjunto de conceitos/codigos deste membro — chave para deduplicar
        # a ontologia compartilhada no agregado (INCLUDE SHARED ONTOLOGY faz
        # membros distintos carregarem os mesmos conceitos).
        concept_keys = set(result.linked_project.ontology_index.keys())
        member_stats.append((alias, result.stats, concept_keys))
        member_results[alias] = result

    if member_errors:
        click.echo(click.style("erro: ao menos um membro falhou na compilacao — link abortado.", fg="red"), err=True)
        raise SystemExit(1)

    # --alpaca ainda nao tem exportador por membro no link step. Avisar em vez
    # de ignorar em silencio — o usuario pedir e nao ver arquivo nenhum e pior
    # que um erro claro. (--xls e exportado por membro mais abaixo, §6.)
    if alpaca_path:
        click.echo(click.style(
            "aviso: --alpaca ainda nao e suportado no passo de linkagem (multiplos projetos) "
            "— nenhum dataset foi gerado. Exporte cada projeto separadamente.",
            fg="yellow",
        ), err=True)

    link_result = link_members(members)
    vr = link_result.validation

    click.echo("")
    _print_diagnostics(vr.errors, "ERROR")
    _print_diagnostics(vr.warnings, "WARNING")

    n_edges = len(link_result.edges)
    n_orphans = len(link_result.orphans)
    if not vr.has_errors():
        entities = sorted({e.entity for e in link_result.edges})
        ent_part = ", ".join(f"'{e}'" for e in entities) if entities else "nenhuma"
        click.echo("")
        click.echo(click.style(f"✓ {len(members)} projetos linkados.", fg="green") +
                   f" {n_edges} aresta(s) resolvida(s) [{ent_part}].")
        if n_orphans:
            click.echo(click.style(f"⚠ {n_orphans} REFERS TO orfao(s) (valor sem IDENTIFIES correspondente).", fg="yellow"))

        if stats:
            click.echo("")
            _print_link_stats(member_stats, n_edges, n_orphans)

    has_errors = vr.has_errors()
    has_warnings = vr.has_warnings()
    exit_code = 1 if has_errors or (strict and has_warnings) else 0

    if (force or exit_code == 0) and not has_errors:
        if json_path:
            payload = {
                "schema_version": "3.1",
                "kind": "link",
                "members": [m.alias for m in members],
                "entity_owners": link_result.entity_owners,
                "links": {
                    "edges": [
                        {
                            "entity": e.entity,
                            "value": e.value,
                            "from": {"member": e.from_member, "bibref": e.from_bibref},
                            "to": {"member": e.to_member, "bibref": e.to_bibref},
                        }
                        for e in link_result.edges
                    ],
                    "orphans": [
                        {"entity": ent, "value": val, "member": mem}
                        for ent, val, mem in link_result.orphans
                    ],
                },
            }
            Path(json_path).write_text(_json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            click.echo(f"  JSON v3.1 agregado -> {json_path}")

        if csv_dir:
            out_dir = Path(csv_dir)
            out_dir.mkdir(parents=True, exist_ok=True)
            links_csv = out_dir / "links.csv"
            with links_csv.open("w", encoding="utf-8", newline="") as fh:
                w = _csv.writer(fh)
                w.writerow(["entity", "value", "from_member", "from_bibref", "to_member", "to_bibref"])
                for e in link_result.edges:
                    w.writerow([e.entity, e.value, e.from_member, e.from_bibref, e.to_member, e.to_bibref])
                for ent, val, mem in link_result.orphans:
                    w.writerow([ent, val, mem, "", "(orphan)", ""])
            click.echo(f"  links.csv -> {links_csv}")

        if xls_path:
            # §6: tabular POR MEMBRO — os SOURCE FIELDS de membros distintos sao
            # incompativeis, entao nao ha TABELA unica coerente. Dois formatos,
            # escolhidos pela extensao do argumento:
            #   --xls saida.xlsx  -> ARQUIVO unico, abas prefixadas por membro
            #                        (lattes_sources, abstracts_sources, ...)
            #                        + aba links. Um anexo so; permite PROCV
            #                        entre abas sem vinculo externo quebradico.
            #   --xls saida/      -> DIRETORIO: um .xlsx por membro + links.xlsx.
            # Em ambos cada membro mantem seu proprio esquema — o que se unifica
            # e o arquivo, nunca as colunas.
            target = Path(xls_path)
            if target.suffix.lower() in (".xlsx", ".xls"):
                collapsed = _export_unified_workbook(
                    link_result, members, member_results, target
                )
                click.echo(f"  workbook unificado -> {target}")
                if collapsed:
                    click.echo(
                        "    abas compartilhadas (identicas em todos os membros): "
                        + ", ".join(collapsed)
                    )
            else:
                target.mkdir(parents=True, exist_ok=True)
                for m in members:
                    res = member_results[m.alias]
                    member_dir = target / m.alias
                    member_dir.mkdir(parents=True, exist_ok=True)
                    member_xlsx = member_dir / f"{m.alias}.xlsx"
                    export_xls(
                        res.linked_project, res.template, member_xlsx,
                        res.bibliography, res.dataset,
                    )
                    click.echo(f"  {m.alias}.xlsx -> {member_xlsx}")

                links_xlsx = target / "links.xlsx"
                _write_links_xlsx(link_result, members, member_results, links_xlsx)
                click.echo(f"  links.xlsx -> {links_xlsx}")

    raise SystemExit(exit_code)


@main.command(cls=_SynesisCommand, epilog=_EPILOG_CHECK)
@click.argument("file", type=click.Path(exists=True))
def check(file: str) -> None:
    """Validate the syntax and integrity of a single Synesis file."""
    try:
        parse_file(Path(file))
        click.echo(click.style("OK", fg="green"))
        raise SystemExit(0)
    except SynesisSyntaxError as exc:
        click.echo(_format_syntax_error(exc), err=True)
        raise SystemExit(1)


@main.command(cls=_SynesisCommand, epilog=_EPILOG_VALIDATE_TEMPLATE)
@click.argument("template", type=click.Path(exists=True))
def validate_template(template: str) -> None:
    """Verify the structure and consistency of a template file."""
    try:
        load_template(Path(template))
        click.echo(click.style("OK", fg="green"))
        raise SystemExit(0)
    except (SynesisSyntaxError, TemplateLoadError) as exc:
        click.echo(str(exc), err=True)
        raise SystemExit(1)


@main.command(cls=_SynesisCommand, epilog=_EPILOG_INIT)
def init() -> None:
    """Create the minimal structure for a new project in the current directory."""
    cwd = Path.cwd()
    project_path = cwd / "project.synp"
    template_path = cwd / "template.synt"
    bibliography_path = cwd / "references.bib"
    annotations_path = cwd / "annotations.syn"
    ontology_path = cwd / "ontology.syno"

    if not project_path.exists():
        project_path.write_text(
            "PROJECT demo\n"
            '    TEMPLATE "template.synt"\n'
            '    INCLUDE BIBLIOGRAPHY "references.bib"\n'
            '    INCLUDE ANNOTATIONS "annotations.syn"\n'
            '    INCLUDE ONTOLOGY "ontology.syno"\n'
            "END PROJECT\n",
            encoding="utf-8",
        )

    if not template_path.exists():
        template_path.write_text(
            "# =================================================================\n"
            "# DATA SOURCES (Interviews, Articles, Field Notes, etc.)\n"
            "# =================================================================\n"
            "\n"
            "SOURCE FIELDS\n"
            "    OPTIONAL description\n"
            "END SOURCE FIELDS\n"
            "\n"
            "FIELD description TYPE TEXT\n"
            "    SCOPE SOURCE\n"
            "    DESCRIPTION General context, summary, or bibliographic details of the data source\n"
            "    GUIDELINES\n"
            "        Summarize the source purpose and context in 1-2 sentences.\n"
            "        Use only information supported by the source.\n"
            "        Do not add analytical interpretation.\n"
            "    END GUIDELINES\n"
            "END FIELD\n"
            "\n"
            "\n"
            "# =================================================================\n"
            "# DATA EXCERPTS & ANALYSIS (Quotes, Memos, and Codes)\n"
            "# =================================================================\n"
            "\n"
            "ITEM FIELDS\n"
            "    REQUIRED citation, note, code\n"
            "END ITEM FIELDS\n"
            "\n"
            "FIELD citation TYPE QUOTATION\n"
            "    SCOPE ITEM\n"
            "    DESCRIPTION Direct quote or selected excerpt from the data source\n"
            "    GUIDELINES\n"
            "        Extract a complete, self-contained excerpt of 1-3 sentences.\n"
            "        Preserve the original wording and punctuation.\n"
            "        Provide enough context for the excerpt to be understood independently.\n"
            "        Do not paraphrase.\n"
            "    END GUIDELINES\n"
            "END FIELD\n"
            "\n"
            "FIELD note TYPE MEMO\n"
            "    SCOPE ITEM\n"
            "    DESCRIPTION Analytical memo recording interpretations, emerging patterns, or causal reasoning\n"
            "    GUIDELINES\n"
            "        Explain the analytical significance of the excerpt in 1-3 sentences.\n"
            "        Identify patterns, mechanisms, or relevant interpretations.\n"
            "        Do not merely restate the citation.\n"
            "        Distinguish textual evidence from your interpretation.\n"
            "    END GUIDELINES\n"
            "END FIELD\n"
            "\n"
            "FIELD code TYPE CODE\n"
            "    SCOPE ITEM\n"
            "    DESCRIPTION Codes or descriptors applied to this specific excerpt (tags)\n"
            "    GUIDELINES\n"
            "        Apply one or more ontology codes directly supported by the excerpt.\n"
            "        Prefer existing codes and avoid redundant synonyms.\n"
            "        Add a new code only for a distinct and analytically relevant concept.\n"
            "        Every code must have a corresponding ONTOLOGY entry.\n"
            "    END GUIDELINES\n"
            "END FIELD\n"
            "\n"
            "\n"
            "# =================================================================\n"
            "# CODEBOOK & THEMES (Coding Framework and Thematic Categories)\n"
            "# =================================================================\n"
            "\n"
            "ONTOLOGY FIELDS\n"
            "    REQUIRED definition, group\n"
            "END ONTOLOGY FIELDS\n"
            "\n"
            "FIELD definition TYPE TEXT\n"
            "    SCOPE ONTOLOGY\n"
            "    DESCRIPTION Clear definition of the code, indicating inclusion/exclusion criteria for when to apply it\n"
            "    GUIDELINES\n"
            "        Define the code in 1-3 sentences.\n"
            "        State when the code should be applied and, when useful, when it should not.\n"
            "        Distinguish it from closely related codes.\n"
            "    END GUIDELINES\n"
            "END FIELD\n"
            "\n"
            "FIELD group TYPE TOPIC\n"
            "    SCOPE ONTOLOGY\n"
            "    DESCRIPTION Broader theme, category, or thematic domain that groups these codes together\n"
            "    GUIDELINES\n"
            "        Assign one broad parent-level thematic category.\n"
            "        Reuse an existing group whenever possible.\n"
            "        Avoid creating a group that applies to only one narrowly defined code.\n"
            "    END GUIDELINES\n"
            "END FIELD\n",
            encoding="utf-8",
        )

    if not bibliography_path.exists():
        bibliography_path.write_text(
            "@article{smith2024,\n"
            "    author = {Smith, Jane},\n"
            "    title = {Understanding Community Resilience},\n"
            "    journal = {Journal of Social Research},\n"
            "    year = {2024},\n"
            "    volume = {12},\n"
            "    pages = {45--67}\n"
            "}\n",
            encoding="utf-8",
        )

    if not annotations_path.exists():
        annotations_path.write_text(
            "SOURCE @smith2024\n"
            "    description: Qualitative study on community resilience strategies in urban contexts.\n"
            "END SOURCE\n"
            "\n"
            "ITEM @smith2024\n"
            "    citation: \"People here look out for each other. When the flood came, nobody waited\n"
            "        for official help — neighbors just organized themselves.\"\n"
            "\n"
            "    note: Participant describes spontaneous collective action as a primary resilience\n"
            "        mechanism, bypassing formal institutions. Suggests strong bonding social capital.\n"
            "\n"
            "    code: Social_Cohesion, Collective_Action\n"
            "END ITEM\n",
            encoding="utf-8",
        )

    if not ontology_path.exists():
        ontology_path.write_text(
            "ONTOLOGY Social_Cohesion\n"
            "    definition: The degree to which community members trust, support, and cooperate\n"
            "        with one another. Applies when participants describe solidarity, mutual aid,\n"
            "        or a shared sense of belonging.\n"
            "    group: Community_Resilience\n"
            "END ONTOLOGY\n"
            "\n"
            "ONTOLOGY Collective_Action\n"
            "    definition: Coordinated efforts by community members to address shared challenges\n"
            "        without formal institutional direction. Applies when groups self-organize in\n"
            "        response to a problem or crisis.\n"
            "    group: Community_Resilience\n"
            "END ONTOLOGY\n",
            encoding="utf-8",
        )

    click.echo(click.style("Basic project initialized.", fg="green"))


def _print_diagnostics(errors: Iterable, severity_label: str, base_dir: Path | None = None) -> None:
    label_color = {"ERROR": "red", "WARNING": "yellow", "INFO": "cyan"}.get(severity_label, "yellow")

    def _fmt_location(loc) -> str:
        try:
            rel = Path(loc.file).relative_to(base_dir) if base_dir else Path(loc.file)
        except ValueError:
            rel = Path(loc.file).name
        return f"{rel}:{loc.line}:{loc.column}"

    formatted = [
        (_fmt_location(err.location), err.to_cli_line())
        for err in errors
    ]
    if not formatted:
        return

    col_width = max(len(loc) for loc, _ in formatted) + 2

    for loc_str, msg in formatted:
        loc_part = click.style(loc_str.ljust(col_width), fg="cyan")
        label_part = click.style(f"[{severity_label}]", fg=label_color, bold=True)
        click.echo(f"{loc_part}{label_part}  {msg}", err=True)


def _print_stats(stats) -> None:
    rows = [
        ("Sources",    stats.source_count),
        ("Items",      stats.item_count),
        ("Ontologies", stats.ontology_count),
        ("Codes",      stats.code_count),
        ("Chains",     stats.chain_count),
    ]
    label_width = max(len(label) for label, _ in rows)
    num_width   = max(len(f"{n:,}".replace(",", ".")) for _, n in rows)

    click.echo(click.style("Estatisticas da Compilacao:", bold=True))
    for label, n in rows:
        formatted_n = f"{n:,}".replace(",", ".")
        click.echo(f"  {label:<{label_width}}  {formatted_n:>{num_width}}")


def _print_link_stats(
    member_stats: list[tuple[str, CompilationStats, set]], n_edges: int, n_orphans: int
) -> None:
    """Estatisticas do passo de linkagem: por membro + agregado.

    Sources/Items/Chains sao PROPRIOS de cada membro -> soma simples. Ja a
    ontologia e compartilhada (INCLUDE SHARED ONTOLOGY): somar os contadores
    daria valor duplicado (74 + 74 = 148 para uma unica ontologia de 74
    conceitos). Por isso ontologia/codigos sao DEDUPLICADOS pela uniao dos
    conjuntos de conceitos, e rotulados como "Shared ..." quando ha sobreposicao.
    """
    # Campos proprios de cada membro (agregam por soma).
    own_fields = [
        ("Sources", "source_count"),
        ("Items",   "item_count"),
        ("Chains",  "chain_count"),
    ]
    # Por-membro tambem exibimos ontologia/codigos (o valor local de cada um).
    per_member_fields = own_fields + [("Ontologies", "ontology_count"), ("Codes", "code_count")]

    alias_width = max(len(alias) for alias, _s, _k in member_stats)

    click.echo(click.style("Estatisticas por membro:", bold=True))
    for alias, s, _keys in member_stats:
        click.echo(f"  {click.style(alias.ljust(alias_width), fg='cyan')}  " + ", ".join(
            f"{label}={getattr(s, attr):,}".replace(",", ".") for label, attr in per_member_fields
        ))

    # Uniao deduplicada dos conceitos/codigos da ontologia compartilhada.
    all_concepts: set = set()
    for _alias, _s, keys in member_stats:
        all_concepts |= keys
    unique_ontology = len(all_concepts)
    summed_ontology = sum(s.ontology_count for _a, s, _k in member_stats)
    shared = summed_ontology > unique_ontology  # ha sobreposicao entre membros

    rows: list[tuple[str, int]] = [
        (label, sum(getattr(s, attr) for _a, s, _k in member_stats))
        for label, attr in own_fields
    ]
    onto_label = "Shared ontology" if shared else "Ontologies"
    code_label = "Shared codes" if shared else "Codes"
    rows.append((onto_label, unique_ontology))
    rows.append((code_label, unique_ontology))
    rows.append(("Edges", n_edges))
    rows.append(("Orphans", n_orphans))

    label_width = max(len(label) for label, _ in rows)
    num_width = max(len(f"{n:,}".replace(",", ".")) for _, n in rows)

    click.echo("")
    click.echo(click.style("Estatisticas agregadas (link step):", bold=True))
    for label, n in rows:
        formatted_n = f"{n:,}".replace(",", ".")
        click.echo(f"  {label:<{label_width}}  {formatted_n:>{num_width}}")


def _format_syntax_error(error: SynesisSyntaxError) -> str:
    return f"{error.location}: [ERROR] {error.message}"


if __name__ == "__main__":
    main()
