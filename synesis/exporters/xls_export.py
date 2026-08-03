"""
xls_export.py - Exportacao XLS do projeto Synesis

Proposito:
    Gerar arquivo XLS unico com multiplas abas, cada aba correspondendo
    a um arquivo CSV que seria gerado pela exportacao CSV.
    Produz rastreabilidade completa para analise em formato Excel.

Componentes principais:
    - export_xls: funcao principal de exportacao

Dependencias criticas:
    - openpyxl: escrita de arquivos Excel (.xlsx)
    - synesis.semantic.linker: LinkedProject consolidado
    - synesis.ast.nodes: TemplateNode para introspeccao

Exemplo de uso:
    from synesis.exporters.xls_export import export_xls
    export_xls(linked, template, Path("saida.xlsx"))

Notas de implementacao:
    - Cada aba corresponde a um CSV (sources, items, ontologies, chains, codes).
    - Apenas gera abas se houver dados e campos relevantes.
    - Todas as abas principais incluem source_file, source_line, source_column.
    - Reutiliza logica do csv_export para consistencia.

Gerado conforme: Especificacao Synesis v1.1
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError(
        "openpyxl nao encontrado. Instale com: pip install openpyxl"
    )

from synesis.ast.nodes import (
    ChainNode,
    FieldType,
    ItemNode,
    Scope,
    SourceNode,
    TemplateNode,
)
from synesis.exporters._helpers import (
    _chain_to_text,
    _get_field_names_for_scope,
    _get_field_names_for_scope_and_types,
    _get_item_field_value,
    _get_ontology_field_value,
    _get_source_field_value,
)
from synesis.semantic.linker import LinkedProject


def build_xls_workbook(
    linked: LinkedProject,
    template: Optional[TemplateNode],
    bibliography: Optional[Dict[str, Any]] = None,
    dataset: Optional[Dict[str, Any]] = None,
    workbook: Optional["Workbook"] = None,
    prefix: str = "",
) -> "Workbook":
    """
    Constroi Workbook Excel em memoria (sem salvar em disco).

    Ideal para manipulacao programatica, streaming ou integracao com APIs.
    O Workbook retornado pode ser salvo posteriormente com wb.save(path).

    Args:
        linked: Projeto vinculado com indices construidos
        template: Template opcional (None = modo legado)
        bibliography: Entradas .bib, para campos ON BIBLIOGRAPHY
        dataset: Registros TOML, para campos ON DATASET
        workbook: Workbook existente onde escrever. Quando None (padrao), um
            novo e criado. Permite acumular varios projetos num unico arquivo
            sem copiar abas entre workbooks — openpyxl nao suporta essa copia
            (`Workbook.copy_worksheet` so opera dentro do mesmo workbook).
        prefix: Prefixo dos nomes de aba (ex.: "lattes" -> "lattes_sources").
            Necessario ao acumular membros, para as abas nao colidirem. O Excel
            limita nomes de aba a 31 caracteres — ver _sheet_name.

    Returns:
        Workbook (openpyxl) com abas:
        - sources: Fontes bibliograficas
        - items: Items anotados
        - ontologies: Conceitos de ontologia
        - chains: Triplas relacionais
        - codes: Frequencia de codigos (modo legado)

    Example:
        >>> wb = build_xls_workbook(linked, template)
        >>> wb.save("output.xlsx")  # Salva quando quiser
        >>> # Ou acumula membros num unico arquivo
        >>> wb = build_xls_workbook(l1, t1, prefix="lattes")
        >>> wb = build_xls_workbook(l2, t2, workbook=wb, prefix="abstracts")
    """
    if workbook is None:
        wb = Workbook()
        # Remove a aba padrao criada automaticamente
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    else:
        wb = workbook

    # Exporta sources se houver campos SOURCE no template
    if template and _has_fields_for_scope(template, Scope.SOURCE):
        _write_sources_sheet(wb, linked, template, bibliography, dataset, prefix)
    elif not template:
        _write_sources_sheet(wb, linked, None, bibliography, dataset, prefix)

    # Exporta items se houver campos ITEM no template
    if template and _has_fields_for_scope(template, Scope.ITEM):
        _write_items_sheet(wb, linked, template, prefix)
    elif not template:
        _write_items_sheet(wb, linked, None, prefix)

    # Exporta ontologies se houver campos ONTOLOGY no template
    if template and _has_fields_for_scope(template, Scope.ONTOLOGY):
        _write_ontologies_sheet(wb, linked, template, prefix)
    elif not template:
        _write_ontologies_sheet(wb, linked, None, prefix)

    # Exporta chains apenas se houver dados de chains no projeto
    has_relations = _detect_chain_relations(linked)
    if _has_chain_data(linked):
        _write_chains_sheet(wb, linked, has_relations, prefix)

    # Exporta codes apenas em modo legado
    if not template and linked.code_usage:
        _write_codes_sheet(wb, linked, prefix)

    # --- Abas derivadas das secoes do JSON que so existiam la ---
    # O XLS era um recorte de 4 visoes enquanto o JSON carrega 9 secoes. As
    # abaixo sao as tabulares por natureza (chave->valor ou lista de tuplas);
    # `project`/`template`/`export_metadata` ficam so no JSON por serem
    # aninhadas e heterogeneas — nao cabem em planilha sem perda de leitura.
    _write_dataset_sheet(wb, linked, template, dataset, prefix)
    _write_code_frequency_sheet(wb, linked, prefix)
    _write_topics_sheet(wb, linked, template, prefix)

    # Se nenhuma aba foi criada, cria uma aba vazia para evitar erro
    if len(wb.sheetnames) == 0:
        wb.create_sheet("Empty")

    return wb


def export_xls(
    linked: LinkedProject,
    template: Optional[TemplateNode],
    output_path: Path,
    bibliography: Optional[Dict[str, Any]] = None,
    dataset: Optional[Dict[str, Any]] = None,
) -> None:
    """
    Exporta projeto Synesis para arquivo XLS unico com multiplas abas.

    Usa build_xls_workbook() para construir os dados e salva em disco.
    Cada aba corresponde a um arquivo CSV que seria gerado pela exportacao CSV.
    """
    if not isinstance(output_path, Path):
        output_path = Path(output_path)

    # Garante extensao .xlsx
    if output_path.suffix.lower() not in ['.xlsx', '.xls']:
        output_path = output_path.with_suffix('.xlsx')

    wb = build_xls_workbook(linked, template, bibliography, dataset)
    wb.save(output_path)


def _sheet_name(base: str, prefix: str = "") -> str:
    """Nome de aba, com prefixo de membro quando houver.

    O Excel limita nomes de aba a 31 caracteres e rejeita o arquivo inteiro se
    algum exceder — entao trunca. Com os aliases reais do corpus os nomes ficam
    bem abaixo do limite (`abstracts_code_frequency` = 24).
    """
    name = f"{prefix}_{base}" if prefix else base
    return name[:31]


def _has_fields_for_scope(template: TemplateNode, scope: Scope) -> bool:
    """Verifica se template define campos para escopo especificado."""
    for spec in template.field_specs.values():
        if spec.scope == scope:
            return True
    return False


def _has_chain_data(linked: LinkedProject) -> bool:
    """Verifica se projeto tem chains."""
    for source in linked.sources.values():
        for item in source.items:
            if item.chains:
                return True
    return False


def _collect_item_bundle_fields(template: TemplateNode) -> set[str]:
    bundles = list(template.bundled_fields.get(Scope.ITEM, []))
    bundles += list(template.optional_bundles.get(Scope.ITEM, []))
    return {name for bundle in bundles for name in bundle}


def _as_list(value: Any) -> List[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _expand_item_rows(
    item: ItemNode,
    field_names: List[str],
    bundle_fields: set[str],
) -> List[Dict[str, Any]]:
    if not bundle_fields:
        return [{name: _get_item_field_value(item, name) for name in field_names}]

    values_by_field: Dict[str, List[Any]] = {}
    max_count = 0
    for name in bundle_fields:
        values = _as_list(_get_item_field_value(item, name))
        values_by_field[name] = values
        if len(values) > max_count:
            max_count = len(values)

    row_count = max(max_count, 1)
    rows: List[Dict[str, Any]] = []
    for idx in range(row_count):
        row: Dict[str, Any] = {}
        for name in field_names:
            if name in bundle_fields:
                values = values_by_field.get(name, [])
                row[name] = values[idx] if idx < len(values) else ""
            else:
                row[name] = _get_item_field_value(item, name)
        rows.append(row)
    return rows


def _write_sources_sheet(
    wb: Workbook,
    linked: LinkedProject,
    template: Optional[TemplateNode],
    bibliography: Optional[Dict[str, Any]] = None,
    dataset: Optional[Dict[str, Any]] = None,
    prefix: str = "",
) -> None:
    sources = list(linked.sources.values())
    if not sources:
        return

    ws = wb.create_sheet(_sheet_name("sources", prefix))

    if template:
        # Usa campos do template
        field_names = _get_field_names_for_scope(template, Scope.SOURCE)
    else:
        # Modo legado: coleta campos dinamicamente
        field_names = _collect_source_fields(sources)

    headers = ["bibref"] + field_names + [
        "source_file",
        "source_line",
        "source_column",
    ]

    # Escreve cabecalho
    ws.append(headers)

    # Escreve dados
    for source in sources:
        location = source.location
        row = [
            source.bibref,
        ]
        for name in field_names:
            row.append(_stringify_value(
                _get_source_field_value(source, name, template, bibliography, dataset)
            ))
        row.extend([
            str(location.file) if location else "",
            location.line if location else "",
            location.column if location else "",
        ])
        ws.append(row)

    # Auto-ajusta largura das colunas
    _auto_size_columns(ws)


def _write_items_sheet(wb: Workbook, linked: LinkedProject, template: Optional[TemplateNode], prefix: str = "") -> None:
    ws = wb.create_sheet(_sheet_name("items", prefix))

    if template:
        # Usa campos do template
        field_names = _get_field_names_for_scope(template, Scope.ITEM)
        headers = ["bibref"] + field_names + [
            "source_file",
            "source_line",
            "source_column",
        ]
    else:
        # Modo legado: hardcoded
        headers = [
            "bibref",
            "quote",
            "codes",
            "note_count",
            "chain_count",
            "source_file",
            "source_line",
            "source_column",
        ]
        field_names = []

    # Escreve cabecalho
    ws.append(headers)

    bundle_fields = _collect_item_bundle_fields(template) if template else set()

    # Escreve dados
    for source in linked.sources.values():
        for item in source.items:
            location = item.location
            if template:
                # Preenche campos do template (expande bundles quando existirem)
                for row_fields in _expand_item_rows(item, field_names, bundle_fields):
                    row = [item.bibref]
                    for name in field_names:
                        row.append(_stringify_value(row_fields.get(name, "")))
                    row.extend([
                        str(location.file) if location else "",
                        location.line if location else "",
                        location.column if location else "",
                    ])
                    ws.append(row)
            else:
                # Modo legado: campos fixos
                row = [item.bibref]
                row.extend([
                    item.quote,
                    ";".join(item.codes),
                    len(item.notes),
                    len(item.chains),
                ])
                row.extend([
                    str(location.file) if location else "",
                    location.line if location else "",
                    location.column if location else "",
                ])
                ws.append(row)

    # Auto-ajusta largura das colunas
    _auto_size_columns(ws)


def _write_ontologies_sheet(wb: Workbook, linked: LinkedProject, template: Optional[TemplateNode], prefix: str = "") -> None:
    if not linked.ontology_index:
        return

    ws = wb.create_sheet(_sheet_name("ontologies", prefix))

    if template:
        # Usa campos do template
        index_fields = _get_field_names_for_scope_and_types(
            template,
            Scope.ITEM,
            {FieldType.CODE, FieldType.CHAIN},
        )
        ontology_fields = _get_field_names_for_scope(template, Scope.ONTOLOGY)
        headers = index_fields + ontology_fields + [
            "source_file",
            "source_line",
            "source_column",
        ]
    else:
        # Modo legado: hardcoded
        headers = [
            "concept",
            "description",
            "topic",
            "aspect",
            "dimension",
            "confidence",
            "source_file",
            "source_line",
            "source_column",
        ]
        field_names = ["topic", "aspect", "dimension", "confidence"]

    # Escreve cabecalho
    ws.append(headers)

    # Escreve dados
    for ontology in linked.ontology_index.values():
        location = ontology.location
        row = []

        if template:
            for name in index_fields:
                row.append(_stringify_value(ontology.concept))
            for name in ontology_fields:
                row.append(_stringify_value(_get_ontology_field_value(ontology, name)))
        else:
            row.extend([
                ontology.concept,
                ontology.description,
            ])
            for name in field_names:
                row.append(_stringify_value(ontology.fields.get(name, "")))

        row.extend([
            str(location.file) if location else "",
            location.line if location else "",
            location.column if location else "",
        ])
        ws.append(row)

    # Auto-ajusta largura das colunas
    _auto_size_columns(ws)


def _write_chains_sheet(wb: Workbook, linked: LinkedProject, has_relations: bool = False, prefix: str = "") -> None:
    ws = wb.create_sheet(_sheet_name("chains", prefix))

    headers = [
        "bibref",
        "from_code",
        "relation",
        "to_code",
        "source_file",
        "source_line",
        "source_column",
    ]

    # Escreve cabecalho
    ws.append(headers)

    # Escreve dados
    for source in linked.sources.values():
        for item in source.items:
            for chain in item.chains:
                for from_code, relation, to_code in chain.to_triples(has_relations=has_relations):
                    location = chain.location
                    row = [
                        item.bibref,
                        from_code,
                        relation,
                        to_code,
                        str(location.file),
                        location.line,
                        location.column,
                    ]
                    ws.append(row)

    # Auto-ajusta largura das colunas
    _auto_size_columns(ws)


def _write_codes_sheet(wb: Workbook, linked: LinkedProject, prefix: str = "") -> None:
    ws = wb.create_sheet(_sheet_name("codes", prefix))

    headers = ["concept", "usage_count", "sources"]

    # Escreve cabecalho
    ws.append(headers)

    # Escreve dados
    for concept, items in linked.code_usage.items():
        sources = sorted({item.bibref for item in items})
        row = [
            concept,
            len(items),
            ";".join(sources),
        ]
        ws.append(row)

    # Auto-ajusta largura das colunas
    _auto_size_columns(ws)


def _write_dataset_sheet(
    wb: Workbook,
    linked: LinkedProject,
    template: Optional[TemplateNode],
    dataset: Optional[Dict[str, Any]],
    prefix: str = "",
) -> None:
    """Aba `dataset`: valores SCOPE SOURCE de origem ON DATASET, por bibref.

    Espelha a secao `dataset` do JSON (separada de `bibliography` para o
    consumidor distinguir a origem). No-op quando o projeto nao usa ON DATASET.
    """
    if not dataset or not template:
        return

    from synesis.parser.dataset_loader import find_record, resolve_path

    ds_specs = [
        (name, spec)
        for name, spec in template.field_specs.items()
        if getattr(spec, "value_origin", "document") == "dataset"
        and spec.scope == Scope.SOURCE
    ]
    if not ds_specs:
        return

    field_names = [name for name, _ in ds_specs]
    ws = wb.create_sheet(_sheet_name("dataset", prefix))
    ws.append(["bibref"] + field_names)

    wrote_row = False
    for bibref, source in linked.sources.items():
        record = find_record(dataset, bibref.lstrip("@"))
        if record is None:
            continue
        row = [source.bibref]
        for name, spec in ds_specs:
            raw = source.fields.get(name)
            if raw is None and spec.dataset_path:
                raw = resolve_path(record, spec.dataset_path)
            # `None` e ausencia: sem o descarte, _stringify_value o converteria
            # na string "None" — indistinguivel de um dado real na planilha.
            row.append("" if raw is None else _stringify_value(raw))
        ws.append(row)
        wrote_row = True

    if not wrote_row:
        wb.remove(ws)
        return
    _auto_size_columns(ws)


def _write_code_frequency_sheet(wb: Workbook, linked: LinkedProject, prefix: str = "") -> None:
    """Aba `code_frequency`: frequencia de uso de cada conceito (indices do JSON)."""
    usage = linked.code_usage
    if not usage:
        return

    ws = wb.create_sheet(_sheet_name("code_frequency", prefix))
    ws.append(["concept", "usage_count", "sources"])
    for concept, items in sorted(usage.items(), key=lambda kv: (-len(kv[1]), kv[0])):
        sources = sorted({item.bibref for item in items})
        ws.append([concept, len(items), ";".join(sources)])
    _auto_size_columns(ws)


def _write_topics_sheet(
    wb: Workbook, linked: LinkedProject, template: Optional[TemplateNode], prefix: str = ""
) -> None:
    """Aba `topics`: conceitos agrupados por topico (indices do JSON), um por linha."""
    topics: Dict[str, List[str]] = {}
    for concept, ontology in linked.ontology_index.items():
        topic = _get_ontology_field_value(ontology, "topic")
        if isinstance(topic, list):
            topic = topic[0] if topic else ""
        topic = str(topic).strip() if topic is not None else ""
        if not topic:
            continue
        topics.setdefault(topic, []).append(concept)

    if not topics:
        return

    ws = wb.create_sheet(_sheet_name("topics", prefix))
    ws.append(["topic", "concept"])
    for topic in sorted(topics):
        for concept in sorted(topics[topic]):
            ws.append([topic, concept])
    _auto_size_columns(ws)


def _collect_source_fields(sources: List[SourceNode]) -> List[str]:
    """Coleta dinamicamente campos de sources (modo legado)."""
    fields = set()
    for source in sources:
        fields.update(source.fields.keys())
    fields.discard("description")
    return sorted(fields)


def _stringify_value(value) -> str:
    """Converte valor para string.

    `None` e ausencia, tambem dentro de listas: sem o descarte, um item nulo
    viraria a string "None" no meio do join — indistinguivel de um dado real.
    ChainNode ganha forma legivel ("a -> REL -> b"); sem isso a celula
    receberia o repr() do dataclass (nodes, locations, WindowsPath).
    """
    if isinstance(value, list):
        return ";".join(_scalar_to_text(v) for v in value if v is not None)
    if value is None:
        return ""
    return _scalar_to_text(value)


def _scalar_to_text(value) -> str:
    """Texto de um valor escalar de celula, tratando nos da AST."""
    if isinstance(value, ChainNode):
        return _chain_to_text(value)
    return str(value)


def _detect_chain_relations(linked: LinkedProject) -> bool:
    """
    Detecta se chains do projeto usam relacoes qualificadas.

    Heuristica: se alguma chain tem numero impar de elementos >= 3,
    provavelmente e qualificada (code -> REL -> code -> REL -> code).
    """
    for source in linked.sources.values():
        for item in source.items:
            for chain in item.chains:
                num_elements = len(chain.nodes)
                # Chain qualificada tem numero impar de elementos >= 3
                if num_elements >= 3 and num_elements % 2 == 1:
                    return True
    return False


def _auto_size_columns(ws) -> None:
    """Auto-ajusta largura das colunas baseado no conteudo."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Limita a 50 caracteres
        ws.column_dimensions[column_letter].width = adjusted_width
