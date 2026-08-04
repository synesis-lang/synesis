"""test_link_step.py - Etapa 2b: passo de linkagem multi-projeto.

Cobre:
  - resolucao de arestas IDENTIFIES/REFERS TO (n:1) sobre fixtures reais;
  - orfao (W083) e quase-casamento (near_match);
  - dono duplicado de entidade (E081);
  - tipos divergentes na entidade (E082);
  - qualificacao de bibref por alias de membro (D10);
  - payload JSON v3.1 do link step.
"""
from pathlib import Path

from synesis.ast.nodes import SourceNode
from synesis.compiler import SynesisCompiler
from synesis.parser.template_loader import load_template_from_string
from synesis.semantic.link_step import Member, link_members

FIXTURES = Path(__file__).parent / "fixtures"


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def _member_from_fixture(alias, sub, synp):
    r = SynesisCompiler(FIXTURES / "T22-Link-Ok" / sub / synp).compile()
    assert not r.has_errors(), [e.CODE for e in r.validation_result.errors]
    return Member(
        alias=alias,
        template=r.template,
        sources=r.linked_project.sources,
        path=FIXTURES / "T22-Link-Ok" / sub / synp,
        bibliography=r.bibliography or {},
    )


def _inline_member(alias, template_src, sources):
    tmpl = load_template_from_string(template_src, f"{alias}.synt")
    src_dict = {bibref: SourceNode(bibref=bibref, fields=fields) for bibref, fields in sources.items()}
    return Member(alias=alias, template=tmpl, sources=src_dict, path=Path(f"{alias}.synp"))


# --------------------------------------------------------------------------
# Resolucao de arestas (fixtures reais, n:1)
# --------------------------------------------------------------------------

def test_link_resolves_n_to_one_edges():
    lattes = _member_from_fixture("lattes", "lattes", "lattes.synp")
    abstracts = _member_from_fixture("abstracts", "abstracts", "abstracts.synp")
    lr = link_members([lattes, abstracts])

    edges = [(e.entity, e.value) for e in lr.edges]
    # artigo_a e artigo_b -> mesmo pesquisador = 2 arestas para o mesmo no
    assert edges.count(("researcher", "3474555741700167")) == 2
    assert lr.validation.errors == []


def test_link_owner_map():
    lattes = _member_from_fixture("lattes", "lattes", "lattes.synp")
    abstracts = _member_from_fixture("abstracts", "abstracts", "abstracts.synp")
    lr = link_members([lattes, abstracts])
    assert lr.entity_owners == {"researcher": "lattes"}


def test_link_orphan_emits_w083():
    lattes = _member_from_fixture("lattes", "lattes", "lattes.synp")
    abstracts = _member_from_fixture("abstracts", "abstracts", "abstracts.synp")
    lr = link_members([lattes, abstracts])
    codes = [w.CODE for w in lr.validation.warnings]
    assert "SYNESIS_W083" in codes
    # o orfao e o pesquisador externo 9999...
    assert ("researcher", "9999999999999999", "abstracts") in lr.orphans


def test_link_qualifies_bibref_by_alias():
    """Formato canonico `alias:@bibref` — o MESMO do synesis-graph (D10)."""
    lattes = _member_from_fixture("lattes", "lattes", "lattes.synp")
    abstracts = _member_from_fixture("abstracts", "abstracts", "abstracts.synp")
    lr = link_members([lattes, abstracts])
    e = next(x for x in lr.edges if x.from_bibref.endswith("artigo_a"))
    assert e.from_bibref == "abstracts:@artigo_a"
    assert e.to_bibref == "lattes:@curriculo_pesq1"


# --------------------------------------------------------------------------
# E081 — dono duplicado de entidade
# --------------------------------------------------------------------------

_OWNER_TMPL = (
    "TEMPLATE {t}\n\n"
    "FIELD k TYPE TEXT\n    SCOPE SOURCE\n    IDENTIFIES researcher\nEND FIELD\n"
)


def test_duplicate_entity_owner_e081():
    a = _inline_member("a", _OWNER_TMPL.format(t="a"), {"x": {"k": "1"}})
    b = _inline_member("b", _OWNER_TMPL.format(t="b"), {"y": {"k": "2"}})
    lr = link_members([a, b])
    codes = [e.CODE for e in lr.validation.errors]
    assert "SYNESIS_E081" in codes


# --------------------------------------------------------------------------
# E082 — tipos divergentes na mesma entidade
# --------------------------------------------------------------------------

def test_type_mismatch_in_linkage_e082():
    owner = _inline_member(
        "owner",
        "TEMPLATE o\n\nFIELD k TYPE TEXT\n    SCOPE SOURCE\n    IDENTIFIES researcher\nEND FIELD\n",
        {"x": {"k": "1"}},
    )
    ref = _inline_member(
        "ref",
        "TEMPLATE r\n\nFIELD k TYPE DATE\n    SCOPE SOURCE\n    REFERS TO researcher\nEND FIELD\n",
        {"y": {"k": "1"}},
    )
    lr = link_members([owner, ref])
    codes = [e.CODE for e in lr.validation.errors]
    assert "SYNESIS_E082" in codes


def test_same_type_no_e082():
    owner = _inline_member(
        "owner",
        "TEMPLATE o\n\nFIELD k TYPE TEXT\n    SCOPE SOURCE\n    IDENTIFIES researcher\nEND FIELD\n",
        {"x": {"k": "1"}},
    )
    ref = _inline_member(
        "ref",
        "TEMPLATE r\n\nFIELD k TYPE TEXT\n    SCOPE SOURCE\n    REFERS TO researcher\nEND FIELD\n",
        {"y": {"k": "1"}},
    )
    lr = link_members([owner, ref])
    assert "SYNESIS_E082" not in [e.CODE for e in lr.validation.errors]


# --------------------------------------------------------------------------
# Quase-casamento (near_match) — detecta sem fundir
# --------------------------------------------------------------------------

def test_near_match_is_warning_not_edge():
    owner = _inline_member(
        "owner",
        "TEMPLATE o\n\nFIELD handle TYPE TEXT\n    SCOPE SOURCE\n    IDENTIFIES account\nEND FIELD\n",
        {"x": {"handle": "@thiagonogueira"}},
    )
    ref = _inline_member(
        "ref",
        "TEMPLATE r\n\nFIELD handle TYPE TEXT\n    SCOPE SOURCE\n    REFERS TO account\nEND FIELD\n",
        {"y": {"handle": "@ThiagoNogueira"}},
    )
    lr = link_members([owner, ref])
    # difere so em caixa: NAO gera aresta (D7), gera warning enriquecido
    assert lr.edges == []
    warn = next(w for w in lr.validation.warnings if w.CODE == "SYNESIS_W083")
    assert warn.near_match == "@thiagonogueira"
    assert "caixa" in warn.to_cli_line().lower()


# --------------------------------------------------------------------------
# CLI: dispatch e payload JSON v3.1
# --------------------------------------------------------------------------

def test_cli_link_produces_v31_json(tmp_path):
    import json

    from click.testing import CliRunner

    from synesis.cli import main

    out_json = tmp_path / "export.json"
    runner = CliRunner()
    result = runner.invoke(main, [
        "compile",
        str(FIXTURES / "T22-Link-Ok" / "lattes" / "lattes.synp"),
        str(FIXTURES / "T22-Link-Ok" / "abstracts" / "abstracts.synp"),
        "--json", str(out_json),
    ])
    assert result.exit_code == 0, result.output
    payload = json.loads(out_json.read_text(encoding="utf-8"))
    assert payload["schema_version"] == "3.1"
    assert payload["kind"] == "link"
    assert payload["entity_owners"] == {"researcher": "lattes"}
    assert len(payload["links"]["edges"]) == 2
    assert len(payload["links"]["orphans"]) == 1


def test_cli_link_stats_shows_per_member_and_aggregate():
    """--stats no link step nao pode ficar mudo (bug real reportado pelo usuario)."""
    from click.testing import CliRunner

    from synesis.cli import main

    runner = CliRunner()
    result = runner.invoke(main, [
        "compile",
        str(FIXTURES / "T22-Link-Ok" / "lattes" / "lattes.synp"),
        str(FIXTURES / "T22-Link-Ok" / "abstracts" / "abstracts.synp"),
        "--stats",
    ])
    assert result.exit_code == 0, result.output
    assert "Estatisticas por membro" in result.output
    assert "lattes" in result.output
    assert "abstracts" in result.output
    assert "Estatisticas agregadas" in result.output
    assert "Edges" in result.output
    assert "Orphans" in result.output


def test_cli_link_without_stats_flag_omits_stats_block():
    from click.testing import CliRunner

    from synesis.cli import main

    runner = CliRunner()
    result = runner.invoke(main, [
        "compile",
        str(FIXTURES / "T22-Link-Ok" / "lattes" / "lattes.synp"),
        str(FIXTURES / "T22-Link-Ok" / "abstracts" / "abstracts.synp"),
    ])
    assert result.exit_code == 0, result.output
    assert "Estatisticas" not in result.output


def _invoke_link_xls(target):
    from click.testing import CliRunner

    from synesis.cli import main

    runner = CliRunner()
    return runner.invoke(main, [
        "compile",
        str(FIXTURES / "T22-Link-Ok" / "lattes" / "lattes.synp"),
        str(FIXTURES / "T22-Link-Ok" / "abstracts" / "abstracts.synp"),
        "--xls", str(target),
    ])


def _assert_links_sheet(ws):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Rotulo legivel de cada lado: sem ele a tabela e so chave estrangeira.
    assert "from_label" in headers
    assert "to_label" in headers
    assert ws.max_row > 1, "nenhuma aresta escrita"


def test_cli_link_xls_dir_exports_per_member_package(tmp_path):
    """--xls SEM extensao no link step: pacote com um .xlsx por membro (§6)."""
    from openpyxl import load_workbook

    out_dir = tmp_path / "pacote"
    result = _invoke_link_xls(out_dir)
    assert result.exit_code == 0, result.output

    assert out_dir.is_dir()
    for alias in ("lattes", "abstracts"):
        assert (out_dir / alias / f"{alias}.xlsx").exists(), f"{alias} nao exportado"

    links_xlsx = out_dir / "links.xlsx"
    assert links_xlsx.exists()
    _assert_links_sheet(load_workbook(links_xlsx)["links"])


def test_cli_link_xls_file_exports_unified_workbook(tmp_path):
    """--xls COM extensao .xlsx: arquivo unico, abas prefixadas por membro.

    Cada membro mantem seu esquema numa aba propria — o que se unifica e o
    arquivo, nunca as colunas (SOURCE FIELDS sao incompativeis entre membros).
    """
    from openpyxl import load_workbook

    out_xls = tmp_path / "unificado.xlsx"
    result = _invoke_link_xls(out_xls)
    assert result.exit_code == 0, result.output

    assert out_xls.is_file()
    assert not (tmp_path / "unificado").exists(), "nao deve criar diretorio"

    wb = load_workbook(out_xls)
    # Abas prefixadas por alias, sem colisao entre membros.
    assert "lattes_sources" in wb.sheetnames
    assert "abstracts_sources" in wb.sheetnames
    assert "links" in wb.sheetnames
    # Limite duro do Excel: nome de aba > 31 chars invalida o arquivo inteiro.
    assert all(len(n) <= 31 for n in wb.sheetnames)

    _assert_links_sheet(wb["links"])


def test_collapse_identical_sheets_moves_shared_to_the_end():
    """Abas fundidas vao para o fim, depois das abas de todos os membros.

    Elas valem para todos, entao nao devem ficar no meio das abas do primeiro
    membro — a posicao sugeriria que sao dele. `links` e criada depois pela
    chamadora, ficando por ultimo.
    """
    from openpyxl import Workbook

    from synesis.cli import _collapse_identical_sheets

    wb = Workbook()
    wb.remove(wb["Sheet"])
    # Ordem de escrita real: bloco de cada membro, na ordem dos argumentos.
    for alias in ("lattes", "abstracts"):
        wb.create_sheet(f"{alias}_sources").append(["bibref", alias])
        wb.create_sheet(f"{alias}_ontologies").append(["concept", "shared"])
        wb.create_sheet(f"{alias}_chains").append(["from", "to", alias])

    collapsed = _collapse_identical_sheets(wb, ["lattes", "abstracts"], ("ontologies",))

    assert collapsed == ["ontologies"]
    # `chains` difere entre membros (coluna com o alias): nao funde.
    assert "lattes_chains" in wb.sheetnames
    assert "abstracts_chains" in wb.sheetnames
    # A fundida e a ultima — `links` entra depois dela.
    assert wb.sheetnames[-1] == "ontologies", wb.sheetnames


def test_cli_link_xls_keeps_per_member_ontology_without_shared_include(tmp_path):
    """Sem INCLUDE SHARED ONTOLOGY, as abas de ontologia NAO sao fundidas.

    A fusao exige a declaracao explicita do autor — nao basta o conteudo
    coincidir, senao uma coincidencia viraria contrato silencioso. As fixtures
    T22 nao declaram SHARED, entao cada membro mantem sua propria aba.
    """
    from openpyxl import load_workbook

    out_xls = tmp_path / "sem_shared.xlsx"
    result = _invoke_link_xls(out_xls)
    assert result.exit_code == 0, result.output

    wb = load_workbook(out_xls)
    assert "ontologies" not in wb.sheetnames, "fundiu sem SHARED ONTOLOGY declarado"
    assert "abas compartilhadas" not in result.output


def test_cli_link_alpaca_warns_instead_of_silently_ignoring(tmp_path):
    from click.testing import CliRunner

    from synesis.cli import main

    out_alpaca = tmp_path / "out.jsonl"
    runner = CliRunner()
    result = runner.invoke(main, [
        "compile",
        str(FIXTURES / "T22-Link-Ok" / "lattes" / "lattes.synp"),
        str(FIXTURES / "T22-Link-Ok" / "abstracts" / "abstracts.synp"),
        "--alpaca", str(out_alpaca),
    ])
    assert result.exit_code == 0, result.output
    assert "aviso" in result.output.lower()
    assert "--alpaca" in result.output
    assert not out_alpaca.exists()


# --------------------------------------------------------------------------
# Dedup de ontologia compartilhada nas estatisticas (--stats)
# --------------------------------------------------------------------------

def _capture_link_stats(member_stats, n_edges, n_orphans):
    import contextlib
    import io

    from synesis.cli import _print_link_stats

    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        _print_link_stats(member_stats, n_edges, n_orphans)
    return buf.getvalue()


def _stats(**kw):
    from synesis.compiler import CompilationStats
    return CompilationStats(**kw)


def test_shared_ontology_stats_are_deduplicated():
    """Dois membros com a MESMA ontologia: agregado mostra o valor unico, nao a soma."""
    shared = {f"c{i}" for i in range(74)}
    member_stats = [
        ("lattes", _stats(source_count=1, item_count=6, ontology_count=74, code_count=74, chain_count=6), set(shared)),
        ("abstracts", _stats(source_count=7, item_count=32, ontology_count=74, code_count=74, chain_count=46), set(shared)),
    ]
    out = _capture_link_stats(member_stats, n_edges=7, n_orphans=0)
    import re
    # rotulo indica compartilhamento
    assert "Ontologia compartilhada" in out
    assert "Shared ontology" in out  # agregado
    # A ontologia sai da tabela por membro: repetir a mesma contagem em cada
    # linha seria ruido (o valor e identico para todos os membros).
    assert "Codes" not in out
    # valor deduplicado (74), nunca a soma (148)
    assert re.search(r"Shared ontology\s+74\b", out)
    assert "148" not in out
    # campos proprios continuam somando (Sources 1+7=8)
    assert re.search(r"Sources\s+8\b", out)


def test_partial_overlap_dedup_counts_union():
    """Sobreposicao parcial: uniao, nao soma nem interseccao."""
    member_stats = [
        ("a", _stats(ontology_count=3, code_count=3), {"x", "y", "z"}),
        ("b", _stats(ontology_count=2, code_count=2), {"z", "w"}),  # z compartilhado
    ]
    out = _capture_link_stats(member_stats, n_edges=0, n_orphans=0)
    # uniao = {x,y,z,w} = 4 (soma seria 5, interseccao seria 1)
    import re
    assert re.search(r"Shared ontology\s+4\b", out)
    assert "Shared ontology" in out  # ha sobreposicao (z), entao rotulo Shared


def test_no_overlap_uses_plain_labels():
    """Sem sobreposicao: rotulo comum, sem nota de compartilhamento."""
    member_stats = [
        ("a", _stats(ontology_count=2, code_count=2), {"x", "y"}),
        ("b", _stats(ontology_count=2, code_count=2), {"w", "z"}),
    ]
    out = _capture_link_stats(member_stats, n_edges=0, n_orphans=0)
    assert "Ontologia:" in out       # sem "compartilhada"
    assert "Shared" not in out
    assert "conceitos unicos" not in out


def test_no_ontology_at_all_omits_ontology_block():
    """Zero conceitos (projetos sem ontologia): o bloco proprio some por
    completo, em vez de imprimir uma linha com zero."""
    member_stats = [
        ("a", _stats(source_count=1), set()),
        ("b", _stats(source_count=3), set()),
    ]
    out = _capture_link_stats(member_stats, n_edges=2, n_orphans=1)
    assert "Shared" not in out
    assert "Ontologia" not in out
    # a tabela por membro e o agregado continuam saindo
    assert "PROJETO" in out
    assert "Sources" in out


# --------------------------------------------------------------------------
# Secao 1 (estrutura) e Secao 2 (resolucao) da saida do link step
# --------------------------------------------------------------------------

_TMPL_OWNER = """TEMPLATE own
SOURCE FIELDS
    REQUIRED pid
END SOURCE FIELDS
ITEM FIELDS
    REQUIRED trecho
END ITEM FIELDS
FIELD pid TYPE TEXT
    SCOPE SOURCE
    IDENTIFIES person
END FIELD
FIELD trecho TYPE QUOTATION
    SCOPE ITEM
END FIELD
"""

_TMPL_REF = """TEMPLATE ref
SOURCE FIELDS
    REQUIRED pid
END SOURCE FIELDS
ITEM FIELDS
    REQUIRED trecho
END ITEM FIELDS
FIELD pid TYPE TEXT
    SCOPE SOURCE
    REFERS TO person
END FIELD
FIELD trecho TYPE QUOTATION
    SCOPE ITEM
END FIELD
"""


def test_link_declarations_derives_from_templates_without_data():
    """A estrutura sai dos templates: uma referencia declarada aparece mesmo
    sem nenhum SOURCE (projeto ainda em coleta)."""
    from synesis.cli import _link_declarations

    members = [
        _inline_member("own", _TMPL_OWNER, {}),
        _inline_member("ref", _TMPL_REF, {}),
    ]
    rows = _link_declarations(members)
    assert rows == [("person", "own", "pid", "ref", "pid")]


def test_link_declarations_marks_entity_without_identifies():
    """REFERS TO sem IDENTIFIES correspondente: colunas de origem vazias."""
    from synesis.cli import _link_declarations

    members = [_inline_member("ref", _TMPL_REF, {})]
    rows = _link_declarations(members)
    assert rows == [("person", "", "", "ref", "pid")]


def _capture_topology(members):
    import io
    from contextlib import redirect_stdout

    from synesis.cli import _print_link_topology

    buf = io.StringIO()
    with redirect_stdout(buf):
        _print_link_topology(members)
    return buf.getvalue()


def test_topology_groups_keywords_over_project_and_field():
    """IDENTIFIES/REFERS TO sao cabecalho de GRUPO sobre (PROJETO, CAMPO) —
    nomear uma coluna `IDENTIFIES` com um projeto embaixo inverteria a relacao
    que o template expressa (a keyword qualifica um FIELD, nao um projeto)."""
    members = [
        _inline_member("own", _TMPL_OWNER, {}),
        _inline_member("ref", _TMPL_REF, {}),
    ]
    out = _capture_topology(members)
    assert "IDENTIFIES" in out
    assert "REFERS TO" in out
    header = next(line for line in out.splitlines() if "ENTITY" in line)
    assert header.split() == ["ENTITY", "PROJETO", "CAMPO", "PROJETO", "CAMPO"]


def test_topology_columns_stay_aligned_with_no_owner_placeholder():
    """'(nenhum)' entra no calculo de largura: e mais longo que muitos nomes
    de projeto e desalinharia a tabela se so os valores reais fossem medidos."""
    members = [_inline_member("r", _TMPL_REF, {})]
    out = _capture_topology(members)
    lines = [ln for ln in out.splitlines() if ln.startswith("  ") and "ENTITY" not in ln]
    data = [ln for ln in lines if "person" in ln]
    assert data, out
    assert "(nenhum)" in data[0]
    # A coluna seguinte comeca depois do placeholder, sem colidir com ele.
    assert data[0].index("(nenhum)") < data[0].index("r ") or " r" in data[0]


def test_cli_link_prints_both_sections_by_default():
    """As duas secoes saem sem --stats: estrutura e util justamente quando
    ainda nao ha dados, que e quando --stats nao tem o que mostrar."""
    from click.testing import CliRunner

    from synesis.cli import main

    runner = CliRunner()
    result = runner.invoke(main, [
        "compile",
        str(FIXTURES / "T22-Link-Ok" / "lattes" / "lattes.synp"),
        str(FIXTURES / "T22-Link-Ok" / "abstracts" / "abstracts.synp"),
    ])
    assert result.exit_code == 0, result.output
    assert "Ligacao entre projetos:" in result.output
    assert "Resolucao das ligacoes:" in result.output
    assert "Estatisticas por membro" not in result.output


def test_cli_link_warns_about_labels_without_edges():
    """Rotulo declarado que nao resolveu nenhuma aresta precisa ser dito: sem
    isso o ✓ le como sucesso pleno mesmo com a ligacao inexistente."""
    from click.testing import CliRunner

    from synesis.cli import main

    runner = CliRunner()
    result = runner.invoke(main, [
        "compile",
        str(FIXTURES / "T22-Link-Ok" / "lattes" / "lattes.synp"),
        str(FIXTURES / "T22-Link-Ok" / "abstracts" / "abstracts.synp"),
    ])
    assert result.exit_code == 0, result.output
    # Fixture resolve 'researcher'; nenhum rotulo deve ser reportado como orfao.
    assert "rotulos sem nenhuma aresta" not in result.output


def test_per_member_table_has_aligned_columns_and_total():
    """A listagem por membro e tabela com TOTAL, nao `k=v` por linha: numeros
    de projetos diferentes precisam ser comparaveis em coluna."""
    member_stats = [
        ("lattes", _stats(source_count=46, item_count=403, chain_count=35), {"x"}),
        ("abstracts", _stats(source_count=7, item_count=32, chain_count=46), {"x"}),
    ]
    out = _capture_link_stats(member_stats, n_edges=7, n_orphans=0)
    header = next(line for line in out.splitlines() if "PROJETO" in line)
    assert header.split() == ["PROJETO", "SOURCES", "ITEMS", "CHAINS"]
    assert "Sources=46" not in out  # formato antigo k=v
    total = next(line for line in out.splitlines() if line.strip().startswith("TOTAL"))
    assert total.split()[1:] == ["53", "435", "81"]


def test_per_member_table_omits_ontology_columns():
    """Ontologia compartilhada nao entra na tabela: seria a mesma contagem
    repetida em todas as linhas."""
    member_stats = [
        ("a", _stats(source_count=1, ontology_count=144), {f"c{i}" for i in range(144)}),
        ("b", _stats(source_count=2, ontology_count=144), {f"c{i}" for i in range(144)}),
    ]
    out = _capture_link_stats(member_stats, n_edges=0, n_orphans=0)
    header = next(line for line in out.splitlines() if "PROJETO" in line)
    assert "ONTOLOG" not in header.upper()
    assert "CODES" not in header.upper()
    # ...e aparece uma unica vez, no bloco proprio
    assert out.count("144") == 2  # bloco de ontologia + linha do agregado
